#!/usr/bin/env python3

import os
import sys
import re
import argparse
from datetime import datetime
import psycopg2
from openpyxl import load_workbook

def connect_to_postgres():
    import psycopg2
    return psycopg2.connect(
        host=os.getenv('PGHOST_2'),
        database=os.getenv('PGDATABASE_2'),
        user=os.getenv('PGUSER_2'),
        password=None,  # psycopg2 uses ~/.pgpass by default
        port=os.getenv('PGPORT_2')
    )

# We expect the following headers in columns A..L:
EXPECTED_HEADERS = [
    "Date",
    "Supply Temp/C",
    "Return Temp/C",
    "Mode",
    "Request",
    "State",
    "Note",
    "Heating",
    "Heating_Group",
    "DateOnly",
    "TimeBlock",
    "TimeBlockState"
]

# Regex for tab (worksheet) name:
TAB_NAME_REGEX = re.compile(r'^(\d{4}-\d{2}-\d{2})\s+(\d{4})-(\d{4})\s+(.+)$')

def parse_tab_name(tab_name):
    """
    Parse the tab name of the format:
      "<YYYY-MM-DD> <HHMM>-<HHMM> <Extra>"
    Return (tab_date, tab_start_time, tab_end_time, tab_extra).
    Raise ValueError if parsing fails.
    """
    match = TAB_NAME_REGEX.match(tab_name.strip())
    if not match:
        raise ValueError(
            f"Worksheet name '{tab_name}' does not match "
            f"the required pattern 'YYYY-MM-DD HHMM-HHMM Extra'"
        )
    tab_date_str, start_str, end_str, tab_extra_str = match.groups()

    # Parse date
    try:
        tab_date = datetime.strptime(tab_date_str, "%Y-%m-%d").date()
    except ValueError:
        raise ValueError(
            f"Cannot parse tab_date '{tab_date_str}' as YYYY-MM-DD in worksheet name '{tab_name}'"
        )

    # Parse start_time
    if len(start_str) != 4:
        raise ValueError(
            f"tab_start_time '{start_str}' in '{tab_name}' must be 4 digits (HHMM)."
        )
    start_hour = int(start_str[:2])
    start_minute = int(start_str[2:])
    try:
        tab_start_time = datetime.strptime(f"{start_hour:02d}:{start_minute:02d}", "%H:%M").time()
    except ValueError:
        raise ValueError(f"Cannot parse start time '{start_str}' in '{tab_name}' as HHMM.")

    # Parse end_time
    if len(end_str) != 4:
        raise ValueError(
            f"tab_end_time '{end_str}' in '{tab_name}' must be 4 digits (HHMM)."
        )
    end_hour = int(end_str[:2])
    end_minute = int(end_str[2:])
    try:
        tab_end_time = datetime.strptime(f"{end_hour:02d}:{end_minute:02d}", "%H:%M").time()
    except ValueError:
        raise ValueError(f"Cannot parse end time '{end_str}' in '{tab_name}' as HHMM.")

    return tab_date, tab_start_time, tab_end_time, tab_extra_str


def verify_header(worksheet):
    """
    Read the first row of the worksheet and ensure it matches EXPECTED_HEADERS.
    If it doesn't match, raise ValueError with a descriptive message.
    """
    header_row = []
    for col_idx in range(1, len(EXPECTED_HEADERS) + 1):
        cell_value = worksheet.cell(row=1, column=col_idx).value
        header_row.append(cell_value if cell_value else "")

    if header_row != EXPECTED_HEADERS:
        raise ValueError(
            f"Header row mismatch in worksheet '{worksheet.title}'.\n"
            f"Expected: {EXPECTED_HEADERS}\n"
            f"Found:    {header_row}"
        )


def parse_time_block(time_block_value):
    """
    Parse the TimeBlock string of the form '<start>-<end>'.
    Return (time_block_start, time_block_end) as time objects.
    Raise ValueError if parsing fails.
    """
    if not time_block_value or '-' not in time_block_value:
        raise ValueError(f"TimeBlock '{time_block_value}' is not in the format 'start-end'.")
    start_str, end_str = time_block_value.split('-', 1)

    def parse_hhmm(s):
        s = s.strip()
        # If there's no colon, but it looks like HHMM, insert a colon
        if ':' not in s and len(s) == 4:
            s = s[:2] + ':' + s[2:]
        return datetime.strptime(s, "%H:%M").time()

    try:
        time_block_start = parse_hhmm(start_str)
        time_block_end = parse_hhmm(end_str)
    except ValueError:
        raise ValueError(f"Cannot parse TimeBlock '{time_block_value}' as 'HH:MM-HH:MM'.")

    return time_block_start, time_block_end


def ensure_table_exists(conn):
    """
    Create the heat_data table if it does not exist.
    Adjust column definitions to match your needs exactly.
    """
    create_table_sql = """
    CREATE TABLE IF NOT EXISTS heat_data (
        -- We'll define columns based on the instructions:
        date            TIMESTAMP WITHOUT TIME ZONE,
        supply          DOUBLE PRECISION,
        return_temp_c   DOUBLE PRECISION,
        mode            VARCHAR,
        request         VARCHAR,
        state           VARCHAR,
        enabled         BOOLEAN,
        note            VARCHAR,
        heating         VARCHAR,
        heating_on      BOOLEAN,
        heating_group   INTEGER,
        date_only       DATE,
        time_block_start TIME,
        time_block_end   TIME,
        timeblockstate   VARCHAR,
        tab_date         DATE,
        tab_start_time   TIME,
        tab_end_time     TIME,
        sheet_name       VARCHAR
    );
    """
    with conn.cursor() as cursor:
        cursor.execute(create_table_sql)
    conn.commit()


def main():
    parser = argparse.ArgumentParser(description="Insert rows from Excel into PostgreSQL.")
    parser.add_argument("--input-file", required=True, help="Path to the Excel .xlsx file.")
    parser.add_argument("--sheet-name", required=True, help="Value to store in each row under column sheet_name.")
    args = parser.parse_args()

    input_file = args.input_file
    global_sheet_name = args.sheet_name

    # Load the workbook
    try:
        wb = load_workbook(filename=input_file, data_only=True)
    except Exception as e:
        print(f"Failed to open the Excel file '{input_file}': {e}")
        sys.exit(1)

    # Connect to PostgreSQL
    try:
        conn = connect_to_postgres()
        conn.autocommit = False
    except Exception as e:
        print(f"Failed to connect to PostgreSQL: {e}")
        sys.exit(1)

    try:
        # Ensure the heat_data table exists before inserting
        ensure_table_exists(conn)

        cursor = conn.cursor()

        # Iterate through each worksheet/tab
        for sheet in wb.worksheets:
            # 1. Parse the tab name
            try:
                tab_date, tab_start_time, tab_end_time, tab_extra = parse_tab_name(sheet.title)
            except ValueError as e:
                print(str(e))
                conn.rollback()
                conn.close()
                sys.exit(1)

            # 2. Verify header
            try:
                verify_header(sheet)
            except ValueError as e:
                print(str(e))
                conn.rollback()
                conn.close()
                sys.exit(1)

            rows_inserted = 0

            # 3. Read data rows (starting from row 2)
            for row_idx in range(2, sheet.max_row + 1):
                row_values = {
                    "Date":             sheet.cell(row=row_idx, column=1).value,
                    "Supply Temp/C":    sheet.cell(row=row_idx, column=2).value,
                    "Return Temp/C":    sheet.cell(row=row_idx, column=3).value,
                    "Mode":             sheet.cell(row=row_idx, column=4).value,
                    "Request":          sheet.cell(row=row_idx, column=5).value,
                    "State":            sheet.cell(row=row_idx, column=6).value,
                    "Note":             sheet.cell(row=row_idx, column=7).value,
                    "Heating":          sheet.cell(row=row_idx, column=8).value,
                    "Heating_Group":    sheet.cell(row=row_idx, column=9).value,
                    "DateOnly":         sheet.cell(row=row_idx, column=10).value,
                    "TimeBlock":        sheet.cell(row=row_idx, column=11).value,
                    "TimeBlockState":   sheet.cell(row=row_idx, column=12).value,
                }

                # Skip row if the Date cell is empty/None
                if row_values["Date"] is None:
                    continue

                # 4. Transform / parse data to match DB columns
                try:
                    # date (timestamp)
                    if isinstance(row_values["Date"], datetime):
                        date_value = row_values["Date"]
                    else:
                        date_value = datetime.strptime(str(row_values["Date"]), "%Y-%m-%d %H:%M:%S")

                    # supply (float)
                    supply = float(row_values["Supply Temp/C"]) if row_values["Supply Temp/C"] is not None else None

                    # return_temp_c (float)
                    return_temp_c = float(row_values["Return Temp/C"]) if row_values["Return Temp/C"] is not None else None

                    # mode (varchar)
                    mode_val = str(row_values["Mode"]) if row_values["Mode"] is not None else ""

                    # request (varchar)
                    request_val = str(row_values["Request"]) if row_values["Request"] is not None else ""

                    # state (varchar)
                    state_val = str(row_values["State"]) if row_values["State"] is not None else ""

                    # enabled (boolean) => true if state == "Enable"
                    enabled_val = (state_val == "Enable")

                    # note (varchar)
                    note_val = str(row_values["Note"]) if row_values["Note"] is not None else ""

                    # heating (varchar)
                    heating_val = str(row_values["Heating"]) if row_values["Heating"] is not None else ""

                    # heating_on (boolean) => true if heating == "On"
                    heating_on_val = (heating_val == "On")

                    # heating_group (int)
                    heating_group_val = None
                    if row_values["Heating_Group"] is not None:
                        heating_group_val = int(row_values["Heating_Group"])

                    # date_only (date)
                    date_only_val = None
                    if isinstance(row_values["DateOnly"], datetime):
                        date_only_val = row_values["DateOnly"].date()
                    elif row_values["DateOnly"] is not None:
                        date_only_val = datetime.strptime(str(row_values["DateOnly"]), "%Y-%m-%d").date()

                    # time_block_start, time_block_end
                    time_block_start, time_block_end = (None, None)
                    tblock_val = row_values["TimeBlock"]
                    if tblock_val:
                        time_block_start, time_block_end = parse_time_block(str(tblock_val))

                    # timeblockstate (varchar)
                    timeblockstate_val = str(row_values["TimeBlockState"]) if row_values["TimeBlockState"] is not None else ""

                except ValueError as ve:
                    print(f"Row {row_idx} in sheet '{sheet.title}' - {ve}")
                    conn.rollback()
                    conn.close()
                    sys.exit(1)

                # 5. Perform the INSERT
                insert_sql = """
                INSERT INTO heat_data (
                    date,
                    supply,
                    return_temp_c,
                    mode,
                    request,
                    state,
                    enabled,
                    note,
                    heating,
                    heating_on,
                    heating_group,
                    date_only,
                    time_block_start,
                    time_block_end,
                    timeblockstate,
                    tab_date,
                    tab_start_time,
                    tab_end_time,
                    sheet_name
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                try:
                    cursor.execute(
                        insert_sql,
                        (
                            date_value,
                            supply,
                            return_temp_c,
                            mode_val,
                            request_val,
                            state_val,
                            enabled_val,
                            note_val,
                            heating_val,
                            heating_on_val,
                            heating_group_val,
                            date_only_val,
                            time_block_start,
                            time_block_end,
                            timeblockstate_val,
                            tab_date,
                            tab_start_time,
                            tab_end_time,
                            global_sheet_name
                        )
                    )
                    rows_inserted += 1
                except Exception as e:
                    print(f"Failed to insert row {row_idx} in sheet '{sheet.title}': {e}")
                    conn.rollback()
                    conn.close()
                    sys.exit(1)

            # Print how many rows we inserted for this tab
            print(f"Tab '{sheet.title}': inserted {rows_inserted} rows.")

        # All good; commit
        conn.commit()

    finally:
        conn.close()


if __name__ == "__main__":
    main()
